from datetime import date, timedelta
from decimal import Decimal, getcontext
from pathlib import Path
from uuid import uuid4

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.db.base import Base
from app.models.bank import Bank
from app.models.loan import Loan
from app.models.rate import Rate
from app.models.transaction import Transaction
from app.services.ledger import compute_ledger
from app.services.reports import build_loan_report


getcontext().prec = 50


@pytest.fixture(scope="session")
def engine():
    eng = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(eng)
    return eng


@pytest.fixture()
def session(engine):
    connection = engine.connect()
    trans = connection.begin()
    Session = sessionmaker(bind=connection, autoflush=False, autocommit=False, future=True)
    s = Session()
    try:
        yield s
    finally:
        s.close()
        trans.rollback()
        connection.close()


def _dec(x) -> Decimal:
    return Decimal(str(x))


def _mk_bank_loan(session, bank_type: str, tenor_months: int, addl_rate_percent: Decimal, placeholder: Decimal):
    bank = Bank(
        name=f"TestBank-{uuid4().hex[:10]}",
        bank_type=bank_type,
        additional_rate=None,
    )
    session.add(bank)
    session.flush()

    loan = Loan(
        bank_id=bank.id,
        name=f"Loan-{uuid4().hex[:10]}",
        kibor_tenor_months=int(tenor_months),
        additional_rate=float(addl_rate_percent),
        kibor_placeholder_rate_percent=float(placeholder),
        max_loan_amount=None,
    )
    session.add(loan)
    session.flush()
    session.commit()
    return bank, loan


def _add_rate(session, bank_id: int, tenor_months: int, effective: date, annual_rate_percent: Decimal):
    r = Rate(
        bank_id=bank_id,
        tenor_months=int(tenor_months),
        effective_date=effective,
        annual_rate_percent=float(annual_rate_percent),
    )
    session.add(r)
    session.commit()
    return r


def _add_tx(session, bank_id: int, loan_id: int, d: date, category: str, amount: Decimal):
    t = Transaction(
        bank_id=bank_id,
        loan_id=loan_id,
        date=d,
        category=category,
        amount=float(amount),
        note=None,
    )
    session.add(t)
    session.commit()
    return t


def _sum_unrounded_daily_markup(principal: Decimal, annual_rate_percent: Decimal, days: int) -> Decimal:
    daily_rate = (annual_rate_percent / Decimal("100")) / Decimal("365")
    daily_markup = principal * daily_rate
    return daily_markup * Decimal(days)


def test_accrued_markup_matches_sum_of_unrounded_daily_markup(session):
    bank, loan = _mk_bank_loan(
        session,
        bank_type="conventional",
        tenor_months=1,
        addl_rate_percent=Decimal("1.234567"),
        placeholder=Decimal("0"),
    )

    start = date(2026, 1, 1)
    end = date(2026, 2, 9)
    days = (end - start).days + 1

    base_rate = Decimal("20.123456")
    _add_rate(session, bank.id, loan.kibor_tenor_months, start, base_rate)

    principal = Decimal("123456789.00")
    _add_tx(session, bank.id, loan.id, start, "principal", principal)

    rows = compute_ledger(session, bank.id, loan.id, start, end)
    assert rows, "Ledger returned no rows"

    last = rows[-1]
    accrued_reported = _dec(last["accrued_markup"])

    annual_total = base_rate + Decimal(str(loan.additional_rate))
    expected_unrounded = _sum_unrounded_daily_markup(principal, annual_total, days)

    diff = abs(accrued_reported - expected_unrounded)
    assert diff <= Decimal("0.005"), (
        f"Accrued markup drifted by {diff}. "
        f"Expected (unrounded accumulation)={expected_unrounded} but got {accrued_reported}. "
        f"This usually indicates rounding during accumulation or float conversion."
    )


def test_conventional_rate_switches_at_month_start(session):
    bank, loan = _mk_bank_loan(
        session,
        bank_type="conventional",
        tenor_months=1,
        addl_rate_percent=Decimal("0.5000"),
        placeholder=Decimal("0"),
    )

    tranche_date = date(2026, 1, 20)
    feb1 = date(2026, 2, 1)

    _add_rate(session, bank.id, loan.kibor_tenor_months, tranche_date, Decimal("10.0000"))
    _add_rate(session, bank.id, loan.kibor_tenor_months, feb1, Decimal("12.0000"))

    _add_tx(session, bank.id, loan.id, tranche_date, "principal", Decimal("1000000.00"))

    rows = compute_ledger(session, bank.id, loan.id, tranche_date, date(2026, 2, 10))
    by_date = {r["date"]: r for r in rows}

    jan31 = date(2026, 1, 31)
    feb2 = date(2026, 2, 2)

    assert jan31 in by_date and feb2 in by_date

    rate_jan31 = _dec(by_date[jan31]["rate_percent"])
    rate_feb2 = _dec(by_date[feb2]["rate_percent"])

    expected_jan31 = Decimal("10.0000") + Decimal(str(loan.additional_rate))
    expected_feb2 = Decimal("12.0000") + Decimal(str(loan.additional_rate))

    assert abs(rate_jan31 - expected_jan31) <= Decimal("0.0005")
    assert abs(rate_feb2 - expected_feb2) <= Decimal("0.0005")


def test_islamic_rate_stays_fixed_after_month_start(session):
    bank, loan = _mk_bank_loan(
        session,
        bank_type="islamic",
        tenor_months=1,
        addl_rate_percent=Decimal("0.5000"),
        placeholder=Decimal("0"),
    )

    tranche_date = date(2026, 1, 20)
    feb1 = date(2026, 2, 1)

    _add_rate(session, bank.id, loan.kibor_tenor_months, tranche_date, Decimal("10.0000"))
    _add_rate(session, bank.id, loan.kibor_tenor_months, feb1, Decimal("12.0000"))

    _add_tx(session, bank.id, loan.id, tranche_date, "principal", Decimal("1000000.00"))

    rows = compute_ledger(session, bank.id, loan.id, tranche_date, date(2026, 2, 10))
    by_date = {r["date"]: r for r in rows}

    feb2 = date(2026, 2, 2)
    assert feb2 in by_date

    rate_feb2 = _dec(by_date[feb2]["rate_percent"])
    expected = Decimal("10.0000") + Decimal(str(loan.additional_rate))
    assert abs(rate_feb2 - expected) <= Decimal("0.0005")


def test_excel_export_contains_more_than_two_decimals_for_daily_markup(tmp_path, session):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    bank, loan = _mk_bank_loan(
        session,
        bank_type="conventional",
        tenor_months=1,
        addl_rate_percent=Decimal("1.234567"),
        placeholder=Decimal("0"),
    )

    start = date(2026, 1, 1)
    end = date(2026, 1, 10)

    base_rate = Decimal("20.123456")
    _add_rate(session, bank.id, loan.kibor_tenor_months, start, base_rate)
    _add_tx(session, bank.id, loan.id, start, "principal", Decimal("123456789.00"))

    out = tmp_path / "loan_report.xlsx"
    build_loan_report(session, bank.id, loan.id, start, end, str(out))

    wb = load_workbook(out)
    ws = wb.active

    daily_markup_cell = ws["C5"].value
    assert isinstance(daily_markup_cell, (int, float)), f"Unexpected Excel cell type: {type(daily_markup_cell)}"

    val = Decimal(str(daily_markup_cell))
    assert val != val.quantize(Decimal("0.01")), (
        f"Excel daily markup appears rounded to 2 decimals ({val}). "
        f"Your requirement says exports must contain unrounded values."
    )